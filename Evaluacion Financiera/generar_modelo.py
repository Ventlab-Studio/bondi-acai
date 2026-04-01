#!/usr/bin/env python3
"""
Bondi Açaí — Modelo de Evaluación Financiera
Genera archivo Excel con análisis financiero completo.

Ejecutar: python generar_modelo.py
Salida:   Bondi_Acai_Evaluacion_Financiera.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter

# ====================================================================
# ESTILOS
# ====================================================================
DARK_BLUE = PatternFill("solid", fgColor="1B3A5C")
MED_BLUE = PatternFill("solid", fgColor="4472C4")
LIGHT_BLUE = PatternFill("solid", fgColor="D6E4F0")
LIGHT_GREEN = PatternFill("solid", fgColor="E2EFDA")
LIGHT_YELLOW = PatternFill("solid", fgColor="FFF2CC")
LIGHT_RED = PatternFill("solid", fgColor="FCE4EC")
LIGHT_GRAY = PatternFill("solid", fgColor="F2F2F2")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD = Font(bold=True, size=11)
BOLD_BIG = Font(bold=True, size=12, color="1B3A5C")
TITLE = Font(bold=True, size=14, color="1B3A5C")
SUBTITLE = Font(bold=True, size=12, color="4472C4")
CUR = '#,##0'          # currency COP sin decimales
PCT = '0.0%'
PCT2 = '0.00%'
NUM1 = '0.0'
THIN_B = Border(bottom=Side('thin', 'CCCCCC'))
THICK_B = Border(top=Side('medium', '1B3A5C'), bottom=Side('medium', '1B3A5C'))
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')

# ====================================================================
# SUPUESTOS Y PARÁMETROS
# ====================================================================
N_MONTHS = 18
DAYS = 28

# --- Inversión inicial ---
INVESTMENT = [
    ("Adecuación del local",        6_000_000,  "Activo fijo",    120),
    ("Equipos esenciales",         15_000_000,  "Activo fijo",     60),
    ("Inventario inicial",          3_000_000,  "Cap. trabajo",     0),
    ("Legal y permisos",            2_500_000,  "Pre-operativo",   12),
    ("Branding y mkt lanzamiento",  2_000_000,  "Intangible",      36),
    ("Capital de trabajo",          6_000_000,  "Cap. trabajo",     0),
    ("Contingencia (10%)",          4_000_000,  "Reserva",          0),
    ("Gastos pre-operativos",       1_500_000,  "Gasto",            0),
]
TOTAL_INV = sum(i[1] for i in INVESTMENT)
STARTING_CASH = 6_000_000 + 4_000_000  # cap. trabajo + contingencia

# --- Escenarios ---
SCENARIOS = {
    "Conservador": {"trans": 30, "ticket": 18_000},
    "Moderado":    {"trans": 45, "ticket": 20_000},
    "Optimista":   {"trans": 65, "ticket": 22_000},
}

# --- Ramp-up ---
RAMP = [0.40, 0.55, 0.65, 0.75, 0.85, 0.92, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0]
GROWTH = 0.03  # mensual post mes 12

def ramp(m):
    return RAMP[m-1] if m <= 12 else (1 + GROWTH) ** (m - 12)

# --- Costos variables (% ingresos) ---
COGS_P   = 0.32
PKG_P    = 0.03
DELIV_P  = 0.015

# --- Costos fijos mensuales ---
NOMINA   = 3_600_000

def founder_comp(m):
    if m <= 6:  return 2_000_000
    if m <= 12: return 2_500_000
    return 3_000_000

SERVICIOS  = 1_000_000
MKT_LAUNCH = 1_300_000   # meses 1-3
MKT_NORMAL = 800_000
CONTAB     = 400_000
SEGUROS    = 200_000
MANT       = 200_000
VARIOS     = 300_000

# --- Depreciación mensual ---
def dep_mensual(m):
    d = 0
    for _, monto, _, meses in INVESTMENT:
        if meses > 0 and m <= meses:
            d += monto / meses
    return d

# --- Impuesto (Régimen Simple) ---
TAX_RATE = 0.035   # 3.5% sobre ingresos brutos

# --- Costo de capital (CAPM) ---
RF         = 0.105    # Tasa libre de riesgo (TES Colombia)
BETA       = 0.90     # Sector alimentos/servicio
MRP        = 0.07     # Prima de riesgo del mercado
SIZE_PREM  = 0.05     # Prima por tamaño / startup
KE_ANUAL   = RF + BETA * MRP + SIZE_PREM   # ~22.3%
KE_MENSUAL = (1 + KE_ANUAL) ** (1/12) - 1  # ~1.69%

# --- Valor terminal ---
TV_MULT = 3.5  # múltiplo EV/EBITDA


# ====================================================================
# CÁLCULOS
# ====================================================================

def compute_pl(sc):
    """Calcula P&L mensual para un escenario."""
    td, tk = sc["trans"], sc["ticket"]
    rows = []
    for m in range(1, N_MONTHS + 1):
        f = ramp(m)
        trans = td * f * DAYS
        rev = trans * tk

        cogs  = rev * COGS_P
        pkg   = rev * PKG_P
        dlv   = rev * DELIV_P
        var_t = cogs + pkg + dlv
        gross = rev - var_t

        fc   = founder_comp(m)
        mktg = MKT_LAUNCH if m <= 3 else MKT_NORMAL
        fix  = NOMINA + fc + SERVICIOS + mktg + CONTAB + SEGUROS + MANT + VARIOS

        ebitda = gross - fix
        dep    = dep_mensual(m)
        ebit   = ebitda - dep
        tax    = rev * TAX_RATE
        ni     = ebit - tax
        fcf    = ebitda - tax   # D&A es no-cash

        rows.append(dict(
            m=m, trans=trans, rev=rev,
            cogs=cogs, pkg=pkg, dlv=dlv, var_t=var_t,
            gross=gross, gross_p=gross/rev if rev else 0,
            nomina=NOMINA, founder=fc, servicios=SERVICIOS,
            mktg=mktg, contab=CONTAB, seguros=SEGUROS,
            mant=MANT, varios=VARIOS, fix=fix,
            ebitda=ebitda, ebitda_p=ebitda/rev if rev else 0,
            dep=dep, ebit=ebit, tax=tax,
            ni=ni, ni_p=ni/rev if rev else 0,
            fcf=fcf,
        ))
    return rows


def irr_newton(cfs, guess=0.02):
    """TIR por método de Newton-Raphson con protección de convergencia."""
    r = guess
    for _ in range(3000):
        try:
            npv = sum(cf / (1+r)**t for t, cf in enumerate(cfs))
            dnpv = sum(-t * cf / (1+r)**(t+1) for t, cf in enumerate(cfs))
        except (OverflowError, ZeroDivisionError):
            return None
        if abs(dnpv) < 1e-15:
            break
        r2 = r - npv / dnpv
        # Limitar a rango razonable [-0.5, 5.0] mensual
        r2 = max(-0.5, min(5.0, r2))
        if abs(r2 - r) < 1e-12:
            return r2
        r = r2
    return r if abs(r) < 5.0 else None


def compute_eval(rows):
    """Indicadores de evaluación financiera."""
    # Flujos: mes 0 = inversión, meses 1-17 = FCF, mes 18 = FCF + valor terminal
    annual_ebitda = rows[-1]["ebitda"] * 12
    tv = annual_ebitda * TV_MULT

    fcfs = [-TOTAL_INV] + [r["fcf"] for r in rows]
    fcfs_tv = list(fcfs)
    fcfs_tv[-1] += tv   # incluir valor terminal en último mes

    # VPN (con valor terminal)
    vpn = sum(cf / (1 + KE_MENSUAL)**t for t, cf in enumerate(fcfs_tv))

    # TIR mensual → anual
    tirm = irr_newton(fcfs_tv)
    tira = (1 + tirm) ** 12 - 1 if tirm is not None else None

    # Payback simple (sin valor terminal)
    cum = 0
    payback = None
    for i, cf in enumerate(fcfs):
        prev = cum
        cum += cf
        if cum >= 0 and payback is None and i > 0:
            payback = (i - 1) + (-prev / cf) if cf else i

    # Payback descontado
    cum_d = 0
    dpayback = None
    for i, cf in enumerate(fcfs):
        dcf = cf / (1 + KE_MENSUAL)**i
        prev = cum_d
        cum_d += dcf
        if cum_d >= 0 and dpayback is None and i > 0:
            dpayback = (i - 1) + (-prev / dcf) if dcf else i

    # RBC (Relación Beneficio-Costo)
    pv_ben = sum(max(0, r["fcf"]) / (1+KE_MENSUAL)**(i+1) for i, r in enumerate(rows))
    pv_ben += tv / (1 + KE_MENSUAL)**N_MONTHS
    pv_cos = TOTAL_INV + sum(max(0, -r["fcf"]) / (1+KE_MENSUAL)**(i+1) for i, r in enumerate(rows))
    rbc = pv_ben / pv_cos if pv_cos else 0

    # Índice de rentabilidad
    ir = (vpn + TOTAL_INV) / TOTAL_INV

    return dict(
        vpn=vpn, tirm=tirm, tira=tira,
        payback=payback, dpayback=dpayback,
        rbc=rbc, ir=ir, tv=tv,
        annual_ebitda=annual_ebitda,
        pv_ben=pv_ben, pv_cos=pv_cos,
    )


# --- Estructuras de inversión ---

def inv_equity_pref(rows, prio=0.70, target=52_000_000, eq_pct=0.30):
    """Opción 1: Equity preferente con recompra."""
    cum = 0; result = []; recovered = False; rec_month = None
    for r in rows:
        dist = max(0, r["ni"])
        share = dist * prio if not recovered else dist * eq_pct
        cum += share
        if cum >= target and not recovered:
            recovered = True
            rec_month = r["m"]
        result.append(dict(m=r["m"], dist=dist, share=share, cum=cum, rec=recovered))
    cfs = [-TOTAL_INV] + [x["share"] for x in result]
    tirm = irr_newton(cfs)
    tira = (1+tirm)**12 - 1 if tirm is not None else None
    return result, tira, rec_month


def inv_loan(rows, rate_m=0.02, grace=3, term=18):
    """Opción 2: Préstamo privado."""
    P = TOTAL_INV
    active = term - grace
    if active > 0 and rate_m > 0:
        pmt = P * rate_m * (1+rate_m)**active / ((1+rate_m)**active - 1)
    else:
        pmt = P / active

    result = []; bal = P; cum = 0
    for m in range(1, term + 1):
        intr = bal * rate_m
        if m <= grace:
            princ = 0; pay = intr
        else:
            princ = pmt - intr; pay = pmt
            bal -= princ
        cum += pay
        result.append(dict(m=m, intr=intr, princ=princ, pay=pay,
                           bal=max(0, bal), cum=cum))
    cfs = [-TOTAL_INV] + [x["pay"] for x in result]
    tirm = irr_newton(cfs)
    tira = (1+tirm)**12 - 1
    return result, tira, pmt, cum


def inv_rev_share(rows, pct=0.12, target=56_000_000):
    """Opción 3: Revenue share."""
    cum = 0; result = []; rec_month = None
    for r in rows:
        if cum < target:
            pay = r["rev"] * pct
            if cum + pay > target:
                pay = target - cum
        else:
            pay = 0
        cum += pay
        if cum >= target and rec_month is None:
            rec_month = r["m"]
        result.append(dict(m=r["m"], pay=pay, cum=cum, rec=cum >= target))
    cfs = [-TOTAL_INV] + [x["pay"] for x in result]
    tirm = irr_newton(cfs)
    tira = (1+tirm)**12 - 1
    return result, tira, rec_month


# ====================================================================
# SENSIBILIDAD
# ====================================================================

def sens_vpn(trans_range, ticket_range):
    """Tabla de sensibilidad VPN bidimensional."""
    matrix = []
    for t in trans_range:
        row = []
        for k in ticket_range:
            try:
                data = compute_pl({"trans": t, "ticket": k})
                ev = compute_eval(data)
                row.append(ev["vpn"])
            except Exception:
                row.append(0)
        matrix.append(row)
    return matrix


# ====================================================================
# HELPERS DE ESCRITURA EXCEL
# ====================================================================

def hdr(ws, row, col, val, span=1, fill=DARK_BLUE, font=WHITE_FONT):
    """Escribe celda de encabezado."""
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill; c.font = font; c.alignment = CENTER
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    return c

def val(ws, row, col, v, fmt=None, bold=False, fill=None, align=None):
    """Escribe celda con formato."""
    c = ws.cell(row=row, column=col, value=v)
    if fmt:  c.number_format = fmt
    if bold: c.font = BOLD
    if fill: c.fill = fill
    if align: c.alignment = align
    else: c.alignment = RIGHT
    return c

def lbl(ws, row, col, v, bold=False, fill=None, indent=0):
    """Escribe etiqueta (alineada izquierda)."""
    c = ws.cell(row=row, column=col, value=("  " * indent + v) if indent else v)
    if bold: c.font = BOLD
    if fill: c.fill = fill
    c.alignment = LEFT
    return c

def title(ws, row, col, v, span=5):
    """Escribe título de sección."""
    c = ws.cell(row=row, column=col, value=v)
    c.font = TITLE
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)

def subtitle(ws, row, col, v, span=5):
    c = ws.cell(row=row, column=col, value=v)
    c.font = SUBTITLE
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)

def set_col_widths(ws, widths):
    """widths: dict {col_letter: width} o {col_num: width}."""
    for k, w in widths.items():
        col_l = k if isinstance(k, str) else get_column_letter(k)
        ws.column_dimensions[col_l].width = w

def total_row(ws, row, col_start, col_end, fill=LIGHT_BLUE):
    """Aplica estilo de fila total."""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD
        cell.fill = fill
        cell.border = THICK_B


# ====================================================================
# HOJAS DEL EXCEL
# ====================================================================

def write_resumen(ws, data, evals):
    """Hoja RESUMEN (dashboard)."""
    ws.sheet_properties.tabColor = "1B3A5C"
    set_col_widths(ws, {1: 35, 2: 20, 3: 20, 4: 20, 5: 5, 6: 25, 7: 20})

    title(ws, 1, 1, "BONDI AÇAÍ — Evaluación Financiera", 7)
    ws.cell(row=2, column=1, value="Modelo generado: Abril 2026").font = Font(italic=True, color="888888")

    # --- Inversión ---
    r = 4
    subtitle(ws, r, 1, "Inversión Requerida"); r += 1
    hdr(ws, r, 1, "Concepto"); hdr(ws, r, 2, "Monto (COP)"); r += 1
    for nm, monto, _, _ in INVESTMENT:
        lbl(ws, r, 1, nm); val(ws, r, 2, monto, CUR); r += 1
    lbl(ws, r, 1, "TOTAL INVERSIÓN", bold=True); val(ws, r, 2, TOTAL_INV, CUR, bold=True)
    total_row(ws, r, 1, 2); r += 2

    # --- WACC ---
    subtitle(ws, r, 1, "Costo del Capital (WACC = Ke)"); r += 1
    params = [
        ("Tasa libre de riesgo (TES)", RF, PCT),
        ("Beta del sector", BETA, NUM1),
        ("Prima de riesgo de mercado", MRP, PCT),
        ("Prima por tamaño/startup", SIZE_PREM, PCT),
        ("Ke anual (Costo del equity)", KE_ANUAL, PCT),
        ("Ke mensual", KE_MENSUAL, PCT2),
    ]
    for label, v, f in params:
        lbl(ws, r, 1, label); val(ws, r, 2, v, f); r += 1
    r += 1

    # --- Indicadores por escenario ---
    subtitle(ws, r, 1, "Indicadores de Evaluación Financiera", 4); r += 1
    hdr(ws, r, 1, "Indicador")
    for i, name in enumerate(SCENARIOS):
        hdr(ws, r, 2+i, name)
    r += 1

    indicators = [
        ("VPN (Valor Presente Neto)", "vpn", CUR),
        ("TIR anual", "tira", PCT),
        ("TIR mensual", "tirm", PCT2),
        ("Payback simple (meses)", "payback", NUM1),
        ("Payback descontado (meses)", "dpayback", NUM1),
        ("Relación Beneficio-Costo (RBC)", "rbc", '0.00'),
        ("Índice de Rentabilidad (IR)", "ir", '0.00'),
        ("Valor Terminal", "tv", CUR),
    ]
    for label, key, fmt in indicators:
        lbl(ws, r, 1, label, bold=True)
        for i, name in enumerate(SCENARIOS):
            v = evals[name][key]
            if v is None: v = "N/A"
            c = val(ws, r, 2+i, v, fmt)
            # Colorear VPN
            if key == "vpn" and isinstance(v, (int, float)):
                c.fill = LIGHT_GREEN if v > 0 else LIGHT_RED
        r += 1
    r += 1

    # --- Resumen P&L anual (Moderado) ---
    subtitle(ws, r, 1, "P&L Resumen — Año 1 (Moderado)", 4); r += 1
    mod = data["Moderado"]
    yr1 = mod[:12]
    totals = {
        "Ingresos": sum(x["rev"] for x in yr1),
        "(-) Costos variables": -sum(x["var_t"] for x in yr1),
        "= Utilidad bruta": sum(x["gross"] for x in yr1),
        "(-) Costos fijos": -sum(x["fix"] for x in yr1),
        "= EBITDA": sum(x["ebitda"] for x in yr1),
        "(-) Depreciación": -sum(x["dep"] for x in yr1),
        "= EBIT": sum(x["ebit"] for x in yr1),
        "(-) Impuesto (Rég. Simple)": -sum(x["tax"] for x in yr1),
        "= Utilidad Neta": sum(x["ni"] for x in yr1),
    }
    hdr(ws, r, 1, "Concepto"); hdr(ws, r, 2, "Año 1 (COP)"); hdr(ws, r, 3, "% Ingresos"); r += 1
    rev_yr = totals["Ingresos"]
    for label, v in totals.items():
        is_total = label.startswith("=")
        lbl(ws, r, 1, label, bold=is_total)
        val(ws, r, 2, v, CUR, bold=is_total)
        val(ws, r, 3, v / rev_yr if rev_yr else 0, PCT, bold=is_total)
        if is_total:
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = LIGHT_BLUE
        r += 1
    r += 1

    # --- Decisión ---
    subtitle(ws, r, 1, "Criterios de Decisión", 4); r += 1
    hdr(ws, r, 1, "Criterio"); hdr(ws, r, 2, "Regla"); hdr(ws, r, 3, "Moderado"); hdr(ws, r, 4, "Resultado"); r += 1
    ev_mod = evals["Moderado"]
    criteria = [
        ("VPN > 0", "Proyecto crea valor", ev_mod["vpn"] > 0),
        ("TIR > WACC (22.3%)", "Rentabilidad supera costo", ev_mod["tira"] > KE_ANUAL if ev_mod["tira"] else False),
        ("RBC > 1", "Beneficios > Costos", ev_mod["rbc"] > 1),
        ("IR > 1", "Inversión rentable", ev_mod["ir"] > 1),
        ("Payback < 18 meses", "Recuperación razonable", ev_mod["payback"] is not None and ev_mod["payback"] < 18),
    ]
    for crit, regla, resultado in criteria:
        lbl(ws, r, 1, crit, bold=True)
        lbl(ws, r, 2, regla)
        v_str = "✓ CUMPLE" if resultado else "✗ NO CUMPLE"
        c = val(ws, r, 3, v_str, align=CENTER)
        c.fill = LIGHT_GREEN if resultado else LIGHT_RED
        c.alignment = CENTER
        r += 1


def write_supuestos(ws):
    """Hoja de supuestos."""
    ws.sheet_properties.tabColor = "4472C4"
    set_col_widths(ws, {1: 40, 2: 20, 3: 15, 4: 40})

    title(ws, 1, 1, "SUPUESTOS DEL MODELO", 4)
    r = 3

    sections = [
        ("Parámetros Generales", [
            ("Horizonte de proyección", f"{N_MONTHS} meses", ""),
            ("Días operativos por mes", str(DAYS), ""),
            ("Régimen tributario", "Régimen Simple de Tributación", ""),
            ("Tasa impositiva (% ingresos brutos)", f"{TAX_RATE:.1%}", "Tarifa consolidada SIMPLE para restaurantes"),
            ("Moneda", "COP (Pesos colombianos)", ""),
        ]),
        ("Escenarios de Ventas", [
            ("Conservador: transacciones/día", "30", "Ticket promedio: $18,000"),
            ("Moderado: transacciones/día", "45", "Ticket promedio: $20,000"),
            ("Optimista: transacciones/día", "65", "Ticket promedio: $22,000"),
        ]),
        ("Ramp-up (% del objetivo)", [
            ("Mes 1", "40%", "Apertura, curva de aprendizaje"),
            ("Mes 2", "55%", "Ganando tracción"),
            ("Mes 3", "65%", "Sampling y boca a boca"),
            ("Mes 4", "75%", "Consolidación"),
            ("Mes 5", "85%", "Base de clientes estable"),
            ("Mes 6", "92%", "Cerca del objetivo"),
            ("Meses 7-12", "100%", "Operación plena"),
            ("Meses 13-18", "100% + 3%/mes", "Crecimiento orgánico"),
        ]),
        ("Costos Variables (% de ingresos)", [
            ("Costo de materia prima (COGS)", f"{COGS_P:.0%}", "Açaí, frutas, granola, base"),
            ("Packaging", f"{PKG_P:.0%}", "Envases, cucharas, bolsas"),
            ("Comisiones delivery", f"{DELIV_P:.1%}", "Rappi y plataformas (~10% de ventas por delivery)"),
        ]),
        ("Costos Fijos Mensuales", [
            ("Nómina empleados (1 FT + 1 PT + prest.)", f"${NOMINA:,.0f}", "Sin gerente (fundadora opera)"),
            ("Remuneración fundadora (meses 1-6)", f"${founder_comp(1):,.0f}", "Salario básico + prestaciones"),
            ("Remuneración fundadora (meses 7-12)", f"${founder_comp(7):,.0f}", "Incremento gradual"),
            ("Remuneración fundadora (meses 13-18)", f"${founder_comp(13):,.0f}", "Estabilización"),
            ("Servicios públicos", f"${SERVICIOS:,.0f}", "Energía, agua, internet"),
            ("Marketing (meses 1-3, lanzamiento)", f"${MKT_LAUNCH:,.0f}", "Campaña de apertura"),
            ("Marketing (mes 4 en adelante)", f"${MKT_NORMAL:,.0f}", "Mantenimiento"),
            ("Contabilidad + software", f"${CONTAB:,.0f}", "Contador + POS"),
            ("Seguros", f"${SEGUROS:,.0f}", ""),
            ("Mantenimiento equipos", f"${MANT:,.0f}", ""),
            ("Varios e imprevistos", f"${VARIOS:,.0f}", ""),
        ]),
        ("Costo del Capital (CAPM)", [
            ("Tasa libre de riesgo (Rf) — TES Colombia", f"{RF:.1%}", "Bonos soberanos Colombia 2026"),
            ("Beta del sector (β)", f"{BETA}", "Alimentos y servicio de comida"),
            ("Prima de riesgo del mercado (Rm - Rf)", f"{MRP:.0%}", "Mercado colombiano"),
            ("Prima por tamaño / startup", f"{SIZE_PREM:.0%}", "Riesgo adicional micro-empresa nueva"),
            ("→ Ke anual = Rf + β(Rm-Rf) + SP", f"{KE_ANUAL:.1%}", "Costo del equity / WACC"),
            ("→ Ke mensual", f"{KE_MENSUAL:.2%}", "Para descontar flujos mensuales"),
        ]),
        ("Valor Terminal", [
            ("Múltiplo EV/EBITDA aplicado", f"{TV_MULT}x", "Rango industria: 2-5x para food service local"),
            ("Método", "EBITDA del mes 18 × 12 × múltiplo", "Going concern"),
        ]),
        ("Ventajas Estructurales (NO incluidas como ingreso)", [
            ("Ahorro por local propio (sin arriendo)", "$3M - $5M COP/mes", "Costo de oportunidad no cargado"),
            ("Ahorro por fundadora como operadora", "Incluido como costo real", "Se paga remuneración modesta"),
            ("Sin máquina de froyo al inicio", "Ahorra $12M-$25M inversión", "Se evalúa en fase 2"),
        ]),
    ]

    for sec_title, items in sections:
        hdr(ws, r, 1, sec_title, span=3); r += 1
        for label, value, note in items:
            lbl(ws, r, 1, label, indent=1)
            val(ws, r, 2, value, align=CENTER)
            lbl(ws, r, 3, note)
            ws.cell(row=r, column=3).font = Font(italic=True, color="666666", size=10)
            r += 1
        r += 1


def write_inversion(ws):
    """Hoja de inversión inicial con depreciación."""
    ws.sheet_properties.tabColor = "4472C4"
    set_col_widths(ws, {1: 30, 2: 18, 3: 15, 4: 15, 5: 18})

    title(ws, 1, 1, "INVERSIÓN INICIAL Y DEPRECIACIÓN", 5)
    r = 3

    headers = ["Categoría", "Monto (COP)", "Tipo", "Vida Útil", "Dep. Mensual (COP)"]
    for i, h in enumerate(headers, 1):
        hdr(ws, r, i, h)
    r += 1

    total_dep = 0
    for nm, monto, tipo, meses in INVESTMENT:
        dep_m = monto / meses if meses > 0 else 0
        total_dep += dep_m
        lbl(ws, r, 1, nm)
        val(ws, r, 2, monto, CUR)
        val(ws, r, 3, tipo, align=CENTER)
        vida = f"{meses} meses" if meses > 0 else "N/A"
        val(ws, r, 4, vida, align=CENTER)
        val(ws, r, 5, dep_m if dep_m > 0 else "—", CUR if dep_m > 0 else None, align=CENTER)
        r += 1

    lbl(ws, r, 1, "TOTAL", bold=True)
    val(ws, r, 2, TOTAL_INV, CUR, bold=True)
    val(ws, r, 5, total_dep, CUR, bold=True)
    total_row(ws, r, 1, 5)
    r += 2

    # Notas
    subtitle(ws, r, 1, "Notas", 5); r += 1
    notas = [
        "• El local es propiedad del padre de la fundadora — no hay arriendo.",
        "• La fundadora opera el negocio directamente — se incluye remuneración modesta.",
        "• No se incluye máquina de frozen yogurt (inversión fase 2, $12-25M adicionales).",
        "• Equipos pueden comprarse usados para reducir inversión 30-50%.",
        "• Capital de trabajo cubre ~2 meses de operación sin ingresos.",
        f"• Depreciación total mes 1-12: ${total_dep:,.0f}/mes.",
        f"• Depreciación total mes 13+: ${dep_mensual(13):,.0f}/mes (legal ya amortizado).",
    ]
    for n in notas:
        lbl(ws, r, 1, n); r += 1


def write_pl(ws, rows, titulo):
    """Hoja de P&L mensual."""
    ws.sheet_properties.tabColor = "70AD47"
    set_col_widths(ws, {1: 32})
    for i in range(2, N_MONTHS + 4):
        ws.column_dimensions[get_column_letter(i)].width = 14

    title(ws, 1, 1, f"ESTADO DE RESULTADOS — {titulo}", N_MONTHS + 2)
    r = 3

    # Encabezados de mes
    hdr(ws, r, 1, "Concepto")
    for m in range(1, N_MONTHS + 1):
        hdr(ws, r, m + 1, f"Mes {m}")
    hdr(ws, r, N_MONTHS + 2, "Año 1")
    hdr(ws, r, N_MONTHS + 3, "Total 18M")
    r += 1

    # Datos
    def write_line(label, key, fmt=CUR, bold=False, fill=None, indent=0, pct_key=None):
        nonlocal r
        lbl(ws, r, 1, label, bold=bold, fill=fill, indent=indent)
        yr1 = sum(rows[i][key] for i in range(12))
        tot = sum(rows[i][key] for i in range(N_MONTHS))
        for m in range(N_MONTHS):
            v = rows[m][key]
            c = val(ws, r, m + 2, v, fmt, bold=bold)
            if fill: c.fill = fill
        c = val(ws, r, N_MONTHS + 2, yr1, fmt, bold=True)
        if fill: c.fill = fill
        c = val(ws, r, N_MONTHS + 3, tot, fmt, bold=True)
        if fill: c.fill = fill
        r += 1

    def write_pct_line(label, key, fill=None):
        nonlocal r
        lbl(ws, r, 1, label, fill=fill, indent=1)
        rev_yr1 = sum(rows[i]["rev"] for i in range(12))
        rev_tot = sum(rows[i]["rev"] for i in range(N_MONTHS))
        for m in range(N_MONTHS):
            v = rows[m][key]
            c = val(ws, r, m + 2, v, PCT)
            if fill: c.fill = fill
            c.font = Font(italic=True, color="666666", size=10)
        yr1_v = sum(rows[i][key.replace("_p", "")] for i in range(12)) / rev_yr1 if rev_yr1 else 0
        tot_v = sum(rows[i][key.replace("_p", "")] for i in range(N_MONTHS)) / rev_tot if rev_tot else 0
        val(ws, r, N_MONTHS + 2, yr1_v, PCT).font = Font(italic=True, color="666666", size=10)
        val(ws, r, N_MONTHS + 3, tot_v, PCT).font = Font(italic=True, color="666666", size=10)
        r += 1

    # Transacciones
    lbl(ws, r, 1, "Transacciones", bold=True)
    for m in range(N_MONTHS):
        val(ws, r, m + 2, rows[m]["trans"], '#,##0')
    val(ws, r, N_MONTHS + 2, sum(rows[i]["trans"] for i in range(12)), '#,##0', bold=True)
    val(ws, r, N_MONTHS + 3, sum(rows[i]["trans"] for i in range(N_MONTHS)), '#,##0', bold=True)
    r += 1

    write_line("INGRESOS", "rev", bold=True, fill=LIGHT_GREEN)
    r += 0  # no extra space

    write_line("(-) Costo materia prima (COGS)", "cogs", indent=1)
    write_line("(-) Packaging", "pkg", indent=1)
    write_line("(-) Comisiones delivery", "dlv", indent=1)
    write_line("= TOTAL COSTOS VARIABLES", "var_t", bold=True)
    r += 0

    write_line("= UTILIDAD BRUTA", "gross", bold=True, fill=LIGHT_BLUE)
    write_pct_line("  Margen bruto %", "gross_p")

    # Costos fijos detallados
    write_line("(-) Nómina empleados", "nomina", indent=1)
    write_line("(-) Remuneración fundadora", "founder", indent=1)
    write_line("(-) Servicios públicos", "servicios", indent=1)
    write_line("(-) Marketing", "mktg", indent=1)
    write_line("(-) Contabilidad + software", "contab", indent=1)
    write_line("(-) Seguros", "seguros", indent=1)
    write_line("(-) Mantenimiento", "mant", indent=1)
    write_line("(-) Varios e imprevistos", "varios", indent=1)
    write_line("= TOTAL COSTOS FIJOS", "fix", bold=True)

    write_line("= EBITDA", "ebitda", bold=True, fill=LIGHT_YELLOW)
    write_pct_line("  Margen EBITDA %", "ebitda_p")

    write_line("(-) Depreciación y amortización", "dep", indent=1)
    write_line("= EBIT (Utilidad Operativa)", "ebit", bold=True)

    write_line("(-) Impuesto Rég. Simple (3.5%)", "tax", indent=1)

    write_line("= UTILIDAD NETA", "ni", bold=True, fill=LIGHT_GREEN)
    write_pct_line("  Margen neto %", "ni_p")

    r += 1
    write_line("FLUJO DE CAJA LIBRE (FCF)", "fcf", bold=True, fill=LIGHT_YELLOW)

    # Saldo de caja acumulado
    lbl(ws, r, 1, "Saldo de caja acumulado", bold=True, fill=LIGHT_BLUE)
    cum_cash = STARTING_CASH
    yr1_final = 0
    for m in range(N_MONTHS):
        cum_cash += rows[m]["fcf"]
        c = val(ws, r, m + 2, cum_cash, CUR, bold=True)
        c.fill = LIGHT_GREEN if cum_cash > 0 else LIGHT_RED
        if m == 11:
            yr1_final = cum_cash
    val(ws, r, N_MONTHS + 2, yr1_final, CUR, bold=True).fill = LIGHT_BLUE
    val(ws, r, N_MONTHS + 3, cum_cash, CUR, bold=True).fill = LIGHT_BLUE


def write_cashflow(ws, data):
    """Hoja de flujo de caja comparativo."""
    ws.sheet_properties.tabColor = "FFC000"
    set_col_widths(ws, {1: 30})
    for i in range(2, N_MONTHS + 4):
        ws.column_dimensions[get_column_letter(i)].width = 14

    title(ws, 1, 1, "FLUJO DE CAJA LIBRE — Comparativo 3 Escenarios", N_MONTHS + 2)
    r = 3

    hdr(ws, r, 1, "")
    for m in range(1, N_MONTHS + 1):
        hdr(ws, r, m + 1, f"Mes {m}")
    hdr(ws, r, N_MONTHS + 2, "Año 1")
    hdr(ws, r, N_MONTHS + 3, "Total 18M")
    r += 1

    for sc_name in SCENARIOS:
        rows = data[sc_name]

        # FCF mensual
        fill = {"Conservador": LIGHT_RED, "Moderado": LIGHT_YELLOW, "Optimista": LIGHT_GREEN}[sc_name]
        lbl(ws, r, 1, f"FCF — {sc_name}", bold=True, fill=fill)
        yr1 = sum(rows[i]["fcf"] for i in range(12))
        tot = sum(rows[i]["fcf"] for i in range(N_MONTHS))
        for m in range(N_MONTHS):
            val(ws, r, m + 2, rows[m]["fcf"], CUR)
        val(ws, r, N_MONTHS + 2, yr1, CUR, bold=True)
        val(ws, r, N_MONTHS + 3, tot, CUR, bold=True)
        r += 1

        # Saldo acumulado
        lbl(ws, r, 1, f"  Saldo acum. — {sc_name}", indent=1)
        cum = STARTING_CASH
        for m in range(N_MONTHS):
            cum += rows[m]["fcf"]
            c = val(ws, r, m + 2, cum, CUR)
            c.fill = LIGHT_GREEN if cum > 0 else LIGHT_RED
            if m == 11: yr1_cum = cum
        val(ws, r, N_MONTHS + 2, yr1_cum, CUR, bold=True)
        val(ws, r, N_MONTHS + 3, cum, CUR, bold=True)
        r += 2

    # Inversión recuperada
    r += 1
    subtitle(ws, r, 1, "Recuperación de la Inversión (sin valor terminal)", N_MONTHS + 2); r += 1
    for sc_name in SCENARIOS:
        rows = data[sc_name]
        lbl(ws, r, 1, f"Acum. FCF neto — {sc_name}", bold=True)
        cum = -TOTAL_INV
        for m in range(N_MONTHS):
            cum += rows[m]["fcf"]
            c = val(ws, r, m + 2, cum, CUR)
            c.fill = LIGHT_GREEN if cum > 0 else LIGHT_RED
        r += 1

    # Gráfico de FCF
    r_chart = r + 2
    chart = LineChart()
    chart.title = "Flujo de Caja Libre Mensual"
    chart.y_axis.title = "COP"
    chart.x_axis.title = "Mes"
    chart.style = 10
    chart.width = 30
    chart.height = 15

    # Escribir datos para gráfico en filas auxiliares
    aux_r = r + 20
    for m in range(1, N_MONTHS + 1):
        ws.cell(row=aux_r, column=m, value=m)
    for i, sc_name in enumerate(SCENARIOS):
        rows = data[sc_name]
        for m in range(N_MONTHS):
            ws.cell(row=aux_r + 1 + i, column=m + 1, value=rows[m]["fcf"])

    cats = Reference(ws, min_col=1, max_col=N_MONTHS, min_row=aux_r)
    for i, sc_name in enumerate(SCENARIOS):
        values = Reference(ws, min_col=1, max_col=N_MONTHS, min_row=aux_r + 1 + i)
        chart.add_data(values, from_rows=True)
        chart.series[i].tx = SeriesLabel(v=sc_name)

    chart.set_categories(cats)
    colors = ["C0504D", "4472C4", "70AD47"]
    for i, c in enumerate(colors):
        if i < len(chart.series):
            chart.series[i].graphicalProperties.line.solidFill = c

    ws.add_chart(chart, f"A{r_chart}")


def write_evaluacion(ws, evals):
    """Hoja de evaluación financiera."""
    ws.sheet_properties.tabColor = "FF0000"
    set_col_widths(ws, {1: 38, 2: 22, 3: 22, 4: 22, 5: 5, 6: 40})

    title(ws, 1, 1, "EVALUACIÓN FINANCIERA DEL PROYECTO", 4)
    r = 3

    # WACC
    subtitle(ws, r, 1, "1. Costo del Capital (WACC)", 4); r += 1
    lbl(ws, r, 1, "Estructura de capital: 100% Equity (sin deuda)"); r += 1
    lbl(ws, r, 1, "WACC = Ke (Costo del Equity) ya que no hay deuda"); r += 1
    r += 1

    hdr(ws, r, 1, "Componente CAPM"); hdr(ws, r, 2, "Valor"); hdr(ws, r, 3, "Fuente"); r += 1
    capm = [
        ("Rf — Tasa libre de riesgo", RF, PCT, "TES Colombia 10 años"),
        ("β — Beta del sector", BETA, '0.00', "Alimentos / food service"),
        ("(Rm - Rf) — Prima de mercado", MRP, PCT, "Mercado accionario Colombia"),
        ("SP — Prima por tamaño", SIZE_PREM, PCT, "Micro-empresa, startup"),
        ("Ke = Rf + β(Rm-Rf) + SP", KE_ANUAL, PCT, "WACC del proyecto"),
    ]
    for label, v, fmt, src in capm:
        lbl(ws, r, 1, label)
        val(ws, r, 2, v, fmt)
        lbl(ws, r, 3, src)
        r += 1
    total_row(ws, r-1, 1, 3, LIGHT_YELLOW)
    r += 2

    # Indicadores
    subtitle(ws, r, 1, "2. Indicadores de Rentabilidad", 4); r += 1
    hdr(ws, r, 1, "Indicador"); hdr(ws, r, 2, "Conservador"); hdr(ws, r, 3, "Moderado"); hdr(ws, r, 4, "Optimista"); r += 1

    indicators = [
        ("VPN (Valor Presente Neto)", "vpn", CUR,
         "VPN > 0 → El proyecto crea valor"),
        ("TIR Anual", "tira", PCT,
         "TIR > WACC (22.3%) → Proyecto rentable"),
        ("TIR Mensual", "tirm", PCT2, ""),
        ("Payback Simple (meses)", "payback", NUM1,
         "Meses para recuperar inversión"),
        ("Payback Descontado (meses)", "dpayback", NUM1,
         "Payback ajustado por valor del dinero"),
        ("Relación Beneficio-Costo (RBC)", "rbc", '0.00',
         "RBC > 1 → Beneficios superan costos"),
        ("Índice de Rentabilidad (IR)", "ir", '0.00',
         "IR > 1 → Inversión genera valor"),
        ("Valor Terminal (VT)", "tv", CUR,
         f"EBITDA anualizado × {TV_MULT}x"),
        ("EBITDA Anualizado (mes 18)", "annual_ebitda", CUR,
         "Base para valor terminal"),
    ]

    for label, key, fmt, interp in indicators:
        lbl(ws, r, 1, label, bold=True)
        for i, nm in enumerate(SCENARIOS):
            v = evals[nm].get(key)
            if v is None: v = "N/A"
            c = val(ws, r, 2 + i, v, fmt)
            if key == "vpn" and isinstance(v, (int, float)):
                c.fill = LIGHT_GREEN if v > 0 else LIGHT_RED
            if key == "tira" and isinstance(v, (int, float)):
                c.fill = LIGHT_GREEN if v > KE_ANUAL else LIGHT_RED
            if key == "rbc" and isinstance(v, (int, float)):
                c.fill = LIGHT_GREEN if v > 1 else LIGHT_RED
        r += 1
        if interp:
            lbl(ws, r, 1, f"  → {interp}", indent=1)
            ws.cell(row=r, column=1).font = Font(italic=True, color="666666", size=10)
            r += 1

    r += 2

    # Interpretación
    subtitle(ws, r, 1, "3. Interpretación de Resultados", 4); r += 1

    ev = evals["Moderado"]
    interpretations = []

    if ev["vpn"] > 0:
        interpretations.append(
            f"✓ VPN POSITIVO (${ev['vpn']:,.0f} COP): El proyecto genera valor por encima del costo de oportunidad del capital. "
            f"Esto significa que invertir $40M en Bondi Açaí es más rentable que cualquier inversión alternativa al {KE_ANUAL:.1%} anual."
        )
    else:
        interpretations.append(
            f"✗ VPN NEGATIVO (${ev['vpn']:,.0f} COP): En el escenario moderado, el proyecto no alcanza a cubrir el costo de oportunidad."
        )

    if ev["tira"] and ev["tira"] > KE_ANUAL:
        interpretations.append(
            f"✓ TIR ({ev['tira']:.1%}) > WACC ({KE_ANUAL:.1%}): La rentabilidad del proyecto supera ampliamente el costo del capital. "
            f"Cada peso invertido genera un retorno superior al mínimo exigido."
        )

    if ev["payback"]:
        interpretations.append(
            f"✓ PAYBACK: La inversión se recupera en ~{ev['payback']:.1f} meses (simple) / "
            f"~{ev['dpayback']:.1f} meses (descontado). "
            f"El benchmark de la industria es 18-36 meses."
        )

    if ev["rbc"] > 1:
        interpretations.append(
            f"✓ RBC = {ev['rbc']:.2f}: Por cada peso de costo, el proyecto genera ${ev['rbc']:.2f} en beneficios."
        )

    for text in interpretations:
        lbl(ws, r, 1, text)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 40
        r += 1

    r += 2
    subtitle(ws, r, 1, "4. Punto de Equilibrio Operativo", 4); r += 1
    # Calcular punto de equilibrio para escenario moderado
    sc = SCENARIOS["Moderado"]
    ticket = sc["ticket"]
    var_pct = COGS_P + PKG_P + DELIV_P  # 36.5%
    mc_unit = ticket * (1 - var_pct)  # margen de contribución por unidad
    fixed_avg = (NOMINA + founder_comp(7) + SERVICIOS + MKT_NORMAL +
                 CONTAB + SEGUROS + MANT + VARIOS)
    tax_unit = ticket * TAX_RATE
    mc_after_tax = mc_unit - tax_unit

    pe_trans = fixed_avg / mc_after_tax
    pe_daily = pe_trans / DAYS
    pe_revenue = pe_trans * ticket

    hdr(ws, r, 1, "Concepto"); hdr(ws, r, 2, "Valor"); r += 1
    pe_data = [
        ("Ticket promedio (moderado)", ticket, CUR),
        ("Costos variables (% del ticket)", var_pct, PCT),
        ("Margen de contribución unitario", mc_unit, CUR),
        ("Impuesto por unidad (3.5%)", tax_unit, CUR),
        ("MC neto por unidad", mc_after_tax, CUR),
        ("Costos fijos mensuales (operación estable)", fixed_avg, CUR),
        ("→ Punto de equilibrio (transacciones/mes)", pe_trans, '#,##0'),
        ("→ Punto de equilibrio (transacciones/día)", pe_daily, '0.0'),
        ("→ Punto de equilibrio (ventas/mes)", pe_revenue, CUR),
    ]
    for label, v, fmt in pe_data:
        lbl(ws, r, 1, label, bold=label.startswith("→"))
        val(ws, r, 2, v, fmt, bold=label.startswith("→"))
        if label.startswith("→"):
            total_row(ws, r, 1, 2, LIGHT_GREEN)
        r += 1


def write_inversionista(ws, mod_rows, eq_res, eq_irr, eq_rec,
                        loan_res, loan_irr, loan_pmt, loan_total,
                        rs_res, rs_irr, rs_rec):
    """Hoja de análisis del inversionista."""
    ws.sheet_properties.tabColor = "7030A0"
    set_col_widths(ws, {1: 35, 2: 20, 3: 20, 4: 20, 5: 20, 6: 20, 7: 20})

    title(ws, 1, 1, "ANÁLISIS DE ESTRUCTURAS CON EL INVERSIONISTA", 6)
    ws.cell(row=2, column=1,
            value="Basado en el escenario MODERADO — Inversión: $40,000,000 COP").font = Font(italic=True, color="666666")
    r = 4

    # === RESUMEN COMPARATIVO ===
    subtitle(ws, r, 1, "Comparativo de Estructuras", 6); r += 1
    hdr(ws, r, 1, "Característica")
    hdr(ws, r, 2, "Opción A\nEquity Preferente", span=2)
    hdr(ws, r, 4, "Opción B\nPréstamo Privado", span=2)
    hdr(ws, r, 6, "Opción C\nRevenue Share")
    r += 1

    comparativo = [
        ("Inversión", "$40M COP", "$40M COP", "$40M COP"),
        ("Participación accionaria", "30% → 10%", "0%", "0%"),
        ("Retorno total al inversionista", "$52M (1.3x)", f"${loan_total:,.0f} ({loan_total/TOTAL_INV:.1f}x)", "$56M (1.4x)"),
        ("Retorno efectivo anual (TIR)", f"{eq_irr:.1%}" if eq_irr else "N/A",
         f"{loan_irr:.1%}" if loan_irr else "N/A",
         f"{rs_irr:.1%}" if rs_irr else "N/A"),
        ("Mes de recuperación total", f"Mes {eq_rec}" if eq_rec else ">18",
         "Mes 18 (plazo fijo)", f"Mes {rs_rec}" if rs_rec else ">18"),
        ("Riesgo para inversionista", "Medio (depende de utilidades)", "Bajo (cuotas fijas)", "Bajo-Medio (depende de ventas)"),
        ("Impacto en flujo de caja", "Variable (% de utilidad)", f"Fijo (${loan_pmt:,.0f}/mes)", "Variable (% de ingresos)"),
        ("Equity residual del inversionista", "10% (recomprable)", "Ninguna", "Ninguna"),
        ("Complejidad legal", "Media", "Baja", "Baja"),
    ]
    for label, a, b, c in comparativo:
        lbl(ws, r, 1, label, bold=True)
        val(ws, r, 2, a, align=CENTER)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        val(ws, r, 4, b, align=CENTER)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        val(ws, r, 6, c, align=CENTER)
        r += 1
    r += 2

    # === OPCIÓN A: EQUITY PREFERENTE ===
    subtitle(ws, r, 1, "OPCIÓN A — Equity con Retorno Preferente y Recompra (RECOMENDADA)", 6); r += 1
    desc_a = [
        "Estructura: El inversionista recibe 30% del equity de la SAS.",
        "Retorno preferente: Recibe 70% de las utilidades netas hasta acumular $52M ($40M + 30% premium).",
        "Post-recuperación: Su participación se reduce automáticamente a 10%.",
        "Opción de recompra: La fundadora puede comprar el 10% restante a valor en libros × 1.2 dentro de 24 meses.",
        "La fundadora recibe 30% de utilidades durante fase preferente, 90% después de la reducción.",
    ]
    for d in desc_a:
        lbl(ws, r, 1, d)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
        r += 1
    r += 1

    # Tabla mensual Opción A
    hdr(ws, r, 1, "Mes"); hdr(ws, r, 2, "Utilidad Neta"); hdr(ws, r, 3, "Distribución Inv. (70%)")
    hdr(ws, r, 4, "Acumulado Inv."); hdr(ws, r, 5, "Distribución Fund. (30%)"); hdr(ws, r, 6, "¿Recuperado?"); r += 1

    for x in eq_res:
        val(ws, r, 1, x["m"], align=CENTER)
        val(ws, r, 2, x["dist"], CUR)
        val(ws, r, 3, x["share"], CUR)
        c = val(ws, r, 4, x["cum"], CUR)
        c.fill = LIGHT_GREEN if x["rec"] else LIGHT_YELLOW
        val(ws, r, 5, x["dist"] * 0.30, CUR)  # fundadora recibe 30%
        val(ws, r, 6, "SÍ" if x["rec"] else "No", align=CENTER)
        r += 1
    r += 2

    # === OPCIÓN B: PRÉSTAMO PRIVADO ===
    subtitle(ws, r, 1, "OPCIÓN B — Préstamo Privado Estructurado", 6); r += 1
    desc_b = [
        "Estructura: Préstamo privado por $40M COP, sin participación accionaria.",
        "Tasa: 2% mensual (24% EA — competitiva para préstamo privado en Colombia).",
        "Gracia: 3 meses (solo intereses), luego cuotas fijas por 15 meses.",
        f"Cuota mensual (mes 4-18): ${loan_pmt:,.0f} COP.",
        f"Total pagado al inversionista: ${loan_total:,.0f} COP.",
        "Ventaja: Sin cesión de equity. Desventaja: Cuotas fijas presionan flujo de caja.",
    ]
    for d in desc_b:
        lbl(ws, r, 1, d)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
        r += 1
    r += 1

    hdr(ws, r, 1, "Mes"); hdr(ws, r, 2, "Interés"); hdr(ws, r, 3, "Capital")
    hdr(ws, r, 4, "Cuota"); hdr(ws, r, 5, "Saldo"); hdr(ws, r, 6, "Acum. Pagado"); r += 1
    for x in loan_res:
        val(ws, r, 1, x["m"], align=CENTER)
        val(ws, r, 2, x["intr"], CUR)
        val(ws, r, 3, x["princ"], CUR)
        val(ws, r, 4, x["pay"], CUR)
        val(ws, r, 5, x["bal"], CUR)
        val(ws, r, 6, x["cum"], CUR)
        r += 1
    r += 2

    # === OPCIÓN C: REVENUE SHARE ===
    subtitle(ws, r, 1, "OPCIÓN C — Revenue Share (Participación en Ingresos)", 6); r += 1
    desc_c = [
        "Estructura: Sin equity. Inversionista recibe 12% de las ventas brutas hasta acumular $56M (1.4x).",
        "Una vez recuperado el total, cesa toda obligación con el inversionista.",
        "Ventaja: Clean exit, sin equity permanente, más simple legalmente.",
        "Desventaja: Pago basado en ingresos (no utilidad), puede ser costoso si hay meses de bajo margen.",
    ]
    for d in desc_c:
        lbl(ws, r, 1, d)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
        r += 1
    r += 1

    hdr(ws, r, 1, "Mes"); hdr(ws, r, 2, "Ventas"); hdr(ws, r, 3, "Pago (12%)")
    hdr(ws, r, 4, "Acum. Pagado"); hdr(ws, r, 5, "¿Recuperado?"); r += 1
    for i, x in enumerate(rs_res):
        val(ws, r, 1, x["m"], align=CENTER)
        val(ws, r, 2, mod_rows[i]["rev"], CUR)
        val(ws, r, 3, x["pay"], CUR)
        c = val(ws, r, 4, x["cum"], CUR)
        c.fill = LIGHT_GREEN if x["rec"] else LIGHT_YELLOW
        val(ws, r, 5, "SÍ" if x["rec"] else "No", align=CENTER)
        r += 1
    r += 2

    # === IMPACTO EN FLUJO DE CAJA ===
    subtitle(ws, r, 1, "Impacto en Flujo de Caja del Negocio — Comparativo", 6); r += 1
    hdr(ws, r, 1, "Mes"); hdr(ws, r, 2, "FCF sin pagos"); hdr(ws, r, 3, "FCF - Opción A")
    hdr(ws, r, 4, "FCF - Opción B"); hdr(ws, r, 5, "FCF - Opción C"); r += 1

    for i in range(N_MONTHS):
        m = i + 1
        fcf = mod_rows[i]["fcf"]
        a_pay = eq_res[i]["share"]
        b_pay = loan_res[i]["pay"] if i < len(loan_res) else 0
        c_pay = rs_res[i]["pay"]

        val(ws, r, 1, m, align=CENTER)
        val(ws, r, 2, fcf, CUR)
        c1 = val(ws, r, 3, fcf - a_pay, CUR)
        c2 = val(ws, r, 4, fcf - b_pay, CUR)
        c3 = val(ws, r, 5, fcf - c_pay, CUR)
        for c in [c1, c2, c3]:
            if isinstance(c.value, (int, float)):
                c.fill = LIGHT_GREEN if c.value > 0 else LIGHT_RED
        r += 1

    r += 2
    subtitle(ws, r, 1, "RECOMENDACIÓN", 6); r += 1
    rec_text = [
        "Se recomienda la OPCIÓN A (Equity Preferente con Recompra) por las siguientes razones:",
        "",
        "1. ALINEACIÓN DE INCENTIVOS: El inversionista gana más cuando el negocio gana más,",
        "   lo que incentiva su apoyo activo (contactos, mentoría, promoción).",
        "",
        "2. FLEXIBILIDAD DE FLUJO: En meses difíciles no hay cuota fija que presione la caja.",
        "   Solo se paga cuando hay utilidad.",
        "",
        "3. SALIDA LIMPIA: La estructura de recompra permite que la fundadora recupere",
        "   el 100% del negocio en un horizonte de 2-3 años.",
        "",
        "4. RETORNO ATRACTIVO PARA EL INVERSIONISTA: 30% de premium sobre el capital + equity",
        "   residual de 10% que puede venderse o mantenerse.",
        "",
        "5. PROTECCIÓN DE LA FUNDADORA: Mantiene 70% del equity desde el inicio",
        "   y llega al 90-100% tras la recuperación del inversionista.",
    ]
    for t in rec_text:
        lbl(ws, r, 1, t)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
        r += 1


def write_sensibilidad(ws):
    """Hoja de sensibilidad."""
    ws.sheet_properties.tabColor = "ED7D31"
    set_col_widths(ws, {1: 30})
    for i in range(2, 12):
        ws.column_dimensions[get_column_letter(i)].width = 18

    title(ws, 1, 1, "ANÁLISIS DE SENSIBILIDAD", 8)
    r = 3

    # Sensibilidad VPN: Transacciones × Ticket
    subtitle(ws, r, 1, "1. VPN vs. Transacciones/día × Ticket Promedio", 8); r += 1
    lbl(ws, r, 1, "VPN (COP) — Celdas verdes: VPN > 0 | Celdas rojas: VPN < 0"); r += 1

    trans_range = [20, 25, 30, 35, 40, 45, 50, 55, 60, 65]
    ticket_range = [15_000, 16_000, 17_000, 18_000, 19_000, 20_000, 21_000, 22_000, 24_000]

    # Header
    hdr(ws, r, 1, "Trans/día ↓  Ticket →")
    for j, tk in enumerate(ticket_range):
        hdr(ws, r, j + 2, f"${tk:,.0f}")
    r += 1

    matrix = sens_vpn(trans_range, ticket_range)
    for i, t in enumerate(trans_range):
        lbl(ws, r, 1, f"{t} trans/día", bold=True)
        for j, vpn_val in enumerate(matrix[i]):
            c = val(ws, r, j + 2, vpn_val, CUR)
            c.fill = LIGHT_GREEN if vpn_val > 0 else LIGHT_RED
        r += 1

    r += 2

    # Sensibilidad unidimensional: COGS %
    subtitle(ws, r, 1, "2. Impacto del COGS % en Utilidad Neta Anual (Moderado)", 8); r += 1
    hdr(ws, r, 1, "COGS %"); hdr(ws, r, 2, "Utilidad Neta Año 1"); hdr(ws, r, 3, "Margen Neto")
    hdr(ws, r, 4, "VPN"); r += 1

    for cogs_test in [0.25, 0.28, 0.30, 0.32, 0.34, 0.36, 0.38, 0.40]:
        # Temporary override
        global COGS_P
        orig = COGS_P
        COGS_P = cogs_test
        rows_t = compute_pl(SCENARIOS["Moderado"])
        ev_t = compute_eval(rows_t)
        ni_yr1 = sum(rows_t[i]["ni"] for i in range(12))
        rev_yr1 = sum(rows_t[i]["rev"] for i in range(12))
        COGS_P = orig

        lbl(ws, r, 1, f"{cogs_test:.0%}", bold=True)
        c1 = val(ws, r, 2, ni_yr1, CUR)
        c2 = val(ws, r, 3, ni_yr1 / rev_yr1 if rev_yr1 else 0, PCT)
        c3 = val(ws, r, 4, ev_t["vpn"], CUR)
        for c in [c1, c3]:
            if isinstance(c.value, (int, float)):
                c.fill = LIGHT_GREEN if c.value > 0 else LIGHT_RED
        r += 1

    r += 2

    # Escenario: ¿Qué pasa si la fundadora no se paga?
    subtitle(ws, r, 1, "3. Escenario: Sin Remuneración de la Fundadora", 8); r += 1
    lbl(ws, r, 1, "Si la fundadora no toma remuneración durante los primeros 12 meses:"); r += 1

    # Calculate without founder comp
    rows_no_fc = []
    sc = SCENARIOS["Moderado"]
    for m in range(1, N_MONTHS + 1):
        f = ramp(m)
        trans = sc["trans"] * f * DAYS
        rev = trans * sc["ticket"]
        var_t = rev * (COGS_P + PKG_P + DELIV_P)
        gross = rev - var_t
        mktg = MKT_LAUNCH if m <= 3 else MKT_NORMAL
        fc_adj = 0 if m <= 12 else founder_comp(m)
        fix = NOMINA + fc_adj + SERVICIOS + mktg + CONTAB + SEGUROS + MANT + VARIOS
        ebitda = gross - fix
        dep = dep_mensual(m)
        ebit = ebitda - dep
        tax = rev * TAX_RATE
        ni = ebit - tax
        fcf = ebitda - tax
        rows_no_fc.append(dict(m=m, rev=rev, ebitda=ebitda, ni=ni, fcf=fcf,
                               var_t=var_t, gross=gross, fix=fix, dep=dep,
                               ebit=ebit, tax=tax,
                               gross_p=gross/rev, ebitda_p=ebitda/rev,
                               ni_p=ni/rev,
                               cogs=rev*COGS_P, pkg=rev*PKG_P, dlv=rev*DELIV_P,
                               trans=trans, nomina=NOMINA, founder=fc_adj,
                               servicios=SERVICIOS, mktg=mktg, contab=CONTAB,
                               seguros=SEGUROS, mant=MANT, varios=VARIOS))

    ev_no_fc = compute_eval(rows_no_fc)
    ev_base = compute_eval(compute_pl(SCENARIOS["Moderado"]))

    hdr(ws, r, 1, "Indicador"); hdr(ws, r, 2, "Con remuneración"); hdr(ws, r, 3, "Sin remuneración"); hdr(ws, r, 4, "Diferencia"); r += 1
    compare = [
        ("Utilidad Neta Año 1",
         sum(compute_pl(SCENARIOS["Moderado"])[i]["ni"] for i in range(12)),
         sum(rows_no_fc[i]["ni"] for i in range(12)), CUR),
        ("VPN", ev_base["vpn"], ev_no_fc["vpn"], CUR),
        ("TIR Anual", ev_base["tira"], ev_no_fc["tira"], PCT),
        ("Payback (meses)", ev_base["payback"], ev_no_fc["payback"], NUM1),
    ]
    for label, v_with, v_without, fmt in compare:
        lbl(ws, r, 1, label, bold=True)
        val(ws, r, 2, v_with, fmt)
        val(ws, r, 3, v_without, fmt)
        if isinstance(v_with, (int, float)) and isinstance(v_without, (int, float)):
            val(ws, r, 4, v_without - v_with, fmt)
        r += 1


def write_pl_all(ws, data):
    """P&L de los 3 escenarios resumido."""
    pass  # Los P&L individuales ya se generan en hojas separadas


# ====================================================================
# MAIN
# ====================================================================

def main():
    # Calcular
    all_data = {}
    all_eval = {}
    for name, sc in SCENARIOS.items():
        all_data[name] = compute_pl(sc)
        all_eval[name] = compute_eval(all_data[name])

    mod = all_data["Moderado"]
    eq_res, eq_irr, eq_rec = inv_equity_pref(mod)
    loan_res, loan_irr, loan_pmt, loan_total = inv_loan(mod)
    rs_res, rs_irr, rs_rec = inv_rev_share(mod)

    # Crear Excel
    wb = Workbook()

    # Resumen (hoja activa)
    ws_res = wb.active
    ws_res.title = "Resumen"
    write_resumen(ws_res, all_data, all_eval)

    # Supuestos
    ws_sup = wb.create_sheet("Supuestos")
    write_supuestos(ws_sup)

    # Inversión
    ws_inv = wb.create_sheet("Inversión Inicial")
    write_inversion(ws_inv)

    # P&L por escenario
    for name in SCENARIOS:
        ws_pl = wb.create_sheet(f"P&L {name}")
        write_pl(ws_pl, all_data[name], f"Escenario {name}")

    # Flujo de caja
    ws_cf = wb.create_sheet("Flujo de Caja")
    write_cashflow(ws_cf, all_data)

    # Evaluación financiera
    ws_ev = wb.create_sheet("Evaluación Financiera")
    write_evaluacion(ws_ev, all_eval)

    # Inversionista
    ws_in = wb.create_sheet("Análisis Inversionista")
    write_inversionista(ws_in, mod, eq_res, eq_irr, eq_rec,
                        loan_res, loan_irr, loan_pmt, loan_total,
                        rs_res, rs_irr, rs_rec)

    # Sensibilidad
    ws_se = wb.create_sheet("Sensibilidad")
    write_sensibilidad(ws_se)

    # Guardar
    outdir = os.path.dirname(os.path.abspath(__file__))
    outpath = os.path.join(outdir, "Bondi_Acai_Evaluacion_Financiera.xlsx")
    wb.save(outpath)
    print(f"\n{'='*60}")
    print(f"  Excel generado exitosamente:")
    print(f"  {outpath}")
    print(f"{'='*60}")
    print(f"\nHojas incluidas:")
    for ws in wb.worksheets:
        print(f"  • {ws.title}")
    print(f"\nParámetros clave:")
    print(f"  Inversión total: ${TOTAL_INV:,.0f} COP")
    print(f"  WACC (Ke): {KE_ANUAL:.1%} anual")
    print(f"  Horizonte: {N_MONTHS} meses")
    print()
    for name in SCENARIOS:
        ev = all_eval[name]
        print(f"  {name}:")
        print(f"    VPN: ${ev['vpn']:,.0f} COP")
        print(f"    TIR: {ev['tira']:.1%} anual")
        pb = f"{ev['payback']:.1f}" if ev['payback'] else ">18"
        print(f"    Payback: {pb} meses")
        print(f"    RBC: {ev['rbc']:.2f}")
        print()


if __name__ == "__main__":
    main()
